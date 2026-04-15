import customtkinter as ctk
from utils.theme import *
from routes.terms import get_all_terms, get_current_term
from routes.assessments import get_assessments
from routes.classes import get_classes
from db.connection import query, query_one


class AnalyticsPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=BG)
        self._terms_data   = []
        self._asmt_data    = []
        self._asmt2_data   = []
        self._classes_data = []
        self._active_tab   = "overview"
        self._build()

    def _build(self):
        self.pack(fill="both", expand=True, padx=20, pady=20)

        heading(self, "Analytics").pack(anchor="w", pady=(0, 12))

        # ── Filter bar ────────────────────────────────────────
        fbar = ctk.CTkFrame(self, fg_color=SURFACE,
                             border_color=BORDER, border_width=1,
                             corner_radius=8)
        fbar.pack(fill="x", pady=(0, 14))
        ff = ctk.CTkFrame(fbar, fg_color="transparent")
        ff.pack(fill="x", padx=14, pady=10)

        muted(ff, "Term:").pack(side="left", padx=(0, 6))
        self._terms_data = get_all_terms()
        term_labels = [f"Term {t['term']}, {t['year']}"
                       for t in self._terms_data]
        current = get_current_term()
        default_t = (f"Term {current['term']}, {current['year']}"
                     if current and term_labels
                     else (term_labels[0] if term_labels else "—"))
        self._term_var = ctk.StringVar(value=default_t)
        ctk.CTkOptionMenu(ff, variable=self._term_var,
                          values=term_labels if term_labels else ["—"],
                          width=150, fg_color=SURFACE,
                          button_color=BORDER, text_color=TEXT,
                          dropdown_fg_color=SURFACE,
                          command=self._on_term_change,
                          ).pack(side="left", padx=(0, 14))

        muted(ff, "Assessment:").pack(side="left", padx=(0, 6))
        self._asmt_var  = ctk.StringVar(value="—")
        self._asmt_data = []
        self._asmt_menu = ctk.CTkOptionMenu(
            ff, variable=self._asmt_var, values=["—"],
            width=190, fg_color=SURFACE, button_color=BORDER,
            text_color=TEXT, dropdown_fg_color=SURFACE,
            command=self._on_filter_change)
        self._asmt_menu.pack(side="left", padx=(0, 14))

        muted(ff, "Class:").pack(side="left", padx=(0, 6))
        self._classes_data = get_classes()
        cls_labels = ["All classes"] + [
            f"{c['name']}{' '+c['stream'] if c.get('stream') else ''}"
            for c in self._classes_data
        ]
        self._class_var = ctk.StringVar(value="All classes")
        ctk.CTkOptionMenu(ff, variable=self._class_var,
                          values=cls_labels, width=170,
                          fg_color=SURFACE, button_color=BORDER,
                          text_color=TEXT, dropdown_fg_color=SURFACE,
                          command=self._on_filter_change,
                          ).pack(side="left", padx=(0, 14))

        # Comparison filter (shown for improvement tabs)
        self._cmp_frame = ctk.CTkFrame(ff, fg_color="transparent")
        muted(self._cmp_frame, "Compare with:").pack(side="left", padx=(0, 6))
        self._asmt2_var  = ctk.StringVar(value="— (none)")
        self._asmt2_data = []
        self._asmt2_menu = ctk.CTkOptionMenu(
            self._cmp_frame, variable=self._asmt2_var,
            values=["— (none)"], width=190,
            fg_color=SURFACE, button_color=BORDER,
            text_color=TEXT, dropdown_fg_color=SURFACE,
            command=self._on_filter_change)
        self._asmt2_menu.pack(side="left")

        # ── Tab bar ───────────────────────────────────────────
        tabs = ctk.CTkFrame(self, fg_color="transparent")
        tabs.pack(fill="x", pady=(0, 12))
        self._tab_btns = {}
        for key, lbl in [
            ("overview",          "School overview"),
            ("subjects",          "Subject performance"),
            ("ranking",           "Exam ranking"),
            ("improved_students", "Most improved students"),
            ("improved_subjects", "Most improved subjects"),
        ]:
            btn = ctk.CTkButton(
                tabs, text=lbl, height=30,
                fg_color=ACCENT if key == "overview" else "transparent",
                text_color="white" if key == "overview" else TEXT_MUTED,
                hover_color=ACCENT_BG, corner_radius=6, font=("", 11),
                command=lambda k=key: self._switch_tab(k),
            )
            btn.pack(side="left", padx=(0, 6))
            self._tab_btns[key] = btn

        # ── Content ───────────────────────────────────────────
        self._content = ctk.CTkFrame(self, fg_color="transparent")
        self._content.pack(fill="both", expand=True)

        # Now safe to load data — _content exists
        self._on_term_change()

    # ── Filter logic ──────────────────────────────────────────
    def _on_term_change(self, _=None):
        t_label = self._term_var.get()
        term = next((t for t in self._terms_data
                     if f"Term {t['term']}, {t['year']}" == t_label), None)
        if not term:
            return
        asmts = get_assessments(term_id=term["id"])
        self._asmt_data = asmts
        labels = [a["name"] for a in asmts]
        self._asmt_menu.configure(values=labels if labels else ["—"])
        self._asmt_var.set(labels[0] if labels else "—")

        all_asmts = get_assessments()
        self._asmt2_data = all_asmts
        all_labels = [a["name"] for a in all_asmts]
        self._asmt2_menu.configure(values=["— (none)"] + all_labels)
        self._asmt2_var.set("— (none)")
        self._on_filter_change()

    def _on_filter_change(self, _=None):
        if self._active_tab in ("improved_students", "improved_subjects"):
            self._cmp_frame.pack(side="left")
        else:
            self._cmp_frame.pack_forget()
        self._render_tab()

    def _switch_tab(self, key):
        self._active_tab = key
        for k, btn in self._tab_btns.items():
            btn.configure(
                fg_color=ACCENT if k == key else "transparent",
                text_color="white" if k == key else TEXT_MUTED)
        self._on_filter_change()

    def _get_filter(self):
        asmt = next((a for a in self._asmt_data
                     if a["name"] == self._asmt_var.get()), None)
        asmt_id = asmt["id"] if asmt else None

        sel = self._class_var.get()
        cls = next((c for c in self._classes_data
                    if f"{c['name']}{' '+c['stream'] if c.get('stream') else ''}"
                    == sel), None)
        class_id = cls["id"] if cls else None

        asmt2 = next((a for a in self._asmt2_data
                      if a["name"] == self._asmt2_var.get()), None)
        asmt2_id = asmt2["id"] if asmt2 else None

        return asmt_id, class_id, asmt2_id

    # ── Tab renderer ─────────────────────────────────────────
    def _render_tab(self):
        for w in self._content.winfo_children():
            w.destroy()
        asmt_id, class_id, asmt2_id = self._get_filter()

        if not asmt_id:
            muted(self._content,
                  "Select a term and assessment to view analytics."
                  ).pack(pady=40)
            return

        scroll = ctk.CTkScrollableFrame(
            self._content, fg_color=BG, corner_radius=0)
        scroll.pack(fill="both", expand=True)

        {
            "overview":          self._tab_overview,
            "subjects":          self._tab_subjects,
            "ranking":           self._tab_ranking,
            "improved_students": self._tab_improved_students,
            "improved_subjects": self._tab_improved_subjects,
        }[self._active_tab](scroll, asmt_id, class_id, asmt2_id)

    # ── Overview ──────────────────────────────────────────────
    def _tab_overview(self, parent, asmt_id, class_id, asmt2_id):
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        if class_id:
            stats = query_one("""
                SELECT COUNT(DISTINCT student_id) AS students,
                       ROUND(AVG(percentage), 1)  AS mean,
                       ROUND(MIN(percentage), 1)  AS min_pct,
                       ROUND(MAX(percentage), 1)  AS max_pct
                FROM marks_new
                WHERE assessment_id=? AND class_id=?
            """, (asmt_id, class_id)) or {}
        else:
            stats = query_one("""
                SELECT COUNT(DISTINCT student_id) AS students,
                       ROUND(AVG(percentage), 1)  AS mean,
                       ROUND(MIN(percentage), 1)  AS min_pct,
                       ROUND(MAX(percentage), 1)  AS max_pct
                FROM marks_new
                WHERE assessment_id=?
            """, (asmt_id,)) or {}

        mean     = stats.get("mean")    or 0
        students = stats.get("students") or 0
        min_pct  = stats.get("min_pct") or 0
        max_pct  = stats.get("max_pct") or 0

        if class_id:
            pass_count = (query_one("""
                SELECT COUNT(*) AS n FROM (
                    SELECT student_id, AVG(percentage) AS avg_pct
                    FROM marks_new WHERE assessment_id=? AND class_id=?
                    GROUP BY student_id
                ) WHERE avg_pct >= 50
            """, (asmt_id, class_id)) or {}).get("n", 0)
        else:
            pass_count = (query_one("""
                SELECT COUNT(*) AS n FROM (
                    SELECT student_id, AVG(percentage) AS avg_pct
                    FROM marks_new WHERE assessment_id=?
                    GROUP BY student_id
                ) WHERE avg_pct >= 50
            """, (asmt_id,)) or {}).get("n", 0)

        pass_rate = round(pass_count / students * 100, 1) if students else 0

        # Stat cards
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=(0, 16))
        row.columnconfigure((0, 1, 2, 3), weight=1, uniform="s")

        for col, (lbl, val, sub, color) in enumerate([
            ("Students analysed", str(students),    "with marks entered",   TEXT_MUTED),
            ("School mean",       f"{mean:.1f}%",   "overall average",      ACCENT),
            ("Pass rate",         f"{pass_rate}%",  "students with ≥ 50%",  SUCCESS if pass_rate >= 50 else DANGER),
            ("Score range",       f"{max_pct:.0f}%", f"Low: {min_pct:.0f}%", TEXT_MUTED),
        ]):
            c = ctk.CTkFrame(row, fg_color=SURFACE,
                             border_color=BORDER, border_width=1,
                             corner_radius=8)
            c.grid(row=0, column=col,
                   padx=(0, 10) if col < 3 else 0, sticky="ew")
            inner = ctk.CTkFrame(c, fg_color="transparent")
            inner.pack(padx=14, pady=12, fill="both")
            muted(inner, lbl, size=11).pack(anchor="w")
            label(inner, val, size=22, weight="bold").pack(
                anchor="w", pady=(2, 0))
            ctk.CTkLabel(inner, text=sub, font=("", 11),
                         text_color=color).pack(anchor="w")

        # Print / export overview button
        ov_btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        ov_btn_row.pack(fill="x", pady=(0, 8))
        _ov_stats = {"students": students, "mean": mean,
                     "pass_rate": pass_rate,
                     "max_pct": max_pct, "min_pct": min_pct}
        ghost_btn(ov_btn_row, "Print / Export overview",
                  command=lambda s=_ov_stats, a=asmt_id, c=class_id:
                      self._print_overview(a, c, s),
                  width=190).pack(side="right")

        # Grade distribution
        if class_id:
            grade_rows = query("""
                SELECT grade, COUNT(*) AS n
                FROM marks_new
                WHERE assessment_id=? AND class_id=?
                GROUP BY grade ORDER BY n DESC
            """, (asmt_id, class_id))
        else:
            grade_rows = query("""
                SELECT grade, COUNT(*) AS n
                FROM marks_new WHERE assessment_id=?
                GROUP BY grade ORDER BY n DESC
            """, (asmt_id,))

        if grade_rows:
            sec = self._section(parent, "Grade distribution")
            self._bar_chart(sec, grade_rows, "grade", "n",
                            show_value=True)

        # Per-class performance
        cls_rows = query("""
            SELECT c.name || COALESCE(' '||c.stream,'') AS cls,
                   ROUND(AVG(m.percentage),1) AS mean,
                   COUNT(DISTINCT m.student_id) AS students
            FROM marks_new m
            JOIN classes c ON m.class_id=c.id
            WHERE m.assessment_id=?
            GROUP BY m.class_id ORDER BY mean DESC
        """, (asmt_id,))

        if cls_rows:
            sec2 = self._section(parent, "Performance by class")
            self._data_table(sec2,
                             ["Class", "Mean %", "Students"],
                             [[r["cls"], f"{r['mean']:.1f}%",
                               str(r["students"])]
                              for r in cls_rows])

    # ── Subject performance ───────────────────────────────────
    def _tab_subjects(self, parent, asmt_id, class_id, asmt2_id):
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        rows = query(f"""
            SELECT s.name AS subject,
                   ROUND(AVG(m.percentage),1) AS mean,
                   ROUND(MIN(m.percentage),1) AS min_pct,
                   ROUND(MAX(m.percentage),1) AS max_pct,
                   COUNT(*) AS entries
            FROM marks_new m
            JOIN subjects s ON m.subject_id=s.id
            WHERE m.assessment_id=? {where}
            GROUP BY m.subject_id ORDER BY mean DESC
        """, params)

        if not rows:
            muted(parent, "No marks found.").pack(pady=24)
            return

        sec = self._section(parent, "Subject performance — ranked by mean")
        self._bar_chart(sec, rows, "subject", "mean", show_value=True)

        tbl = [[str(i+1), r["subject"], f"{r['mean']:.1f}%",
                 f"{r['min_pct']:.1f}%", f"{r['max_pct']:.1f}%",
                 str(r["entries"])]
                for i, r in enumerate(rows)]
        sec2 = self._section(parent, "Detail")
        self._export_row(sec2,
            ["#", "Subject", "Mean %", "Min %", "Max %", "Count"], tbl)
        self._data_table(sec2,
            ["#", "Subject", "Mean %", "Min %", "Max %", "Count"], tbl,
            col_widths=[40, 280, 90, 90, 90, 70])

    # ── Exam ranking ──────────────────────────────────────────
    def _tab_ranking(self, parent, asmt_id, class_id, asmt2_id):
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        rows = query(f"""
            SELECT st.full_name, st.admission_number,
                   c.name || COALESCE(' '||c.stream,'') AS cls,
                   ROUND(AVG(m.percentage),1) AS mean
            FROM marks_new m
            JOIN students st ON m.student_id=st.id
            JOIN classes  c  ON m.class_id=c.id
            WHERE m.assessment_id=? {where}
            GROUP BY m.student_id ORDER BY mean DESC
        """, params)

        if not rows:
            muted(parent, "No marks found.").pack(pady=24)
            return

        from utils.grading import grade_from_percentage, detect_curriculum
        show_class = class_id is None

        data = []
        pos  = 1
        for i, r in enumerate(rows):
            if i > 0 and r["mean"] < rows[i-1]["mean"]:
                pos = i + 1
            # Get grade for this student's mean
            cls_name = r["cls"].split()[0] if r["cls"] else ""
            curriculum = detect_curriculum(cls_name)
            grade, _ = grade_from_percentage(r["mean"], curriculum)
            row_data = [str(pos), r["full_name"],
                        r["admission_number"]]
            if show_class:
                row_data.append(r["cls"])
            row_data += [f"{r['mean']:.1f}", grade]
            data.append(row_data)

        headers = ["#", "Full name", "Adm. No."]
        if show_class:
            headers.append("Class")
        headers += ["Mean", "Grade"]

        widths = [40, 220, 120]
        if show_class:
            widths.append(150)
        widths += [70, 70]

        sec = self._section(parent,
                            f"Exam ranking — {len(data)} student(s)")
        self._export_row(sec, headers, data)
        self._data_table(sec, headers, data,
                         highlight_top=3, col_widths=widths)

    # ── Most improved students ────────────────────────────────
    def _tab_improved_students(self, parent, asmt_id,
                                class_id, asmt2_id):
        if not asmt2_id:
            info = ctk.CTkFrame(parent, fg_color=ACCENT_BG,
                                corner_radius=8)
            info.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(info,
                         text="Select a baseline assessment in 'Compare with' above.",
                         font=("", 12), text_color=ACCENT,
                         ).pack(padx=14, pady=10, anchor="w")
            return

        where  = "AND m.class_id=?" if class_id else ""
        p1 = (asmt_id,  class_id) if class_id else (asmt_id,)
        p2 = (asmt2_id, class_id) if class_id else (asmt2_id,)

        curr = {r["student_id"]: r["mean"] for r in query(f"""
            SELECT student_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY student_id
        """, p1)}

        prev = {r["student_id"]: r["mean"] for r in query(f"""
            SELECT student_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY student_id
        """, p2)}

        improvements = []
        for sid, cur_m in curr.items():
            if sid in prev:
                diff = round(cur_m - prev[sid], 1)
                improvements.append((sid, prev[sid], cur_m, diff))

        if not improvements:
            muted(parent, "No students found in both assessments."
                  ).pack(pady=24)
            return

        improvements.sort(key=lambda x: x[3], reverse=True)

        id_map = {r["id"]: r for r in query(
            "SELECT id, full_name, admission_number FROM students")}

        data = []
        for sid, prev_m, cur_m, diff in improvements:
            s = id_map.get(sid, {})
            arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "—")
            data.append([s.get("full_name", "—"),
                          s.get("admission_number", "—"),
                          f"{prev_m:.1f}%", f"{cur_m:.1f}%",
                          f"{arrow} {abs(diff):.1f}%"])

        hdrs = ["Full name", "Adm. No.", "Previous", "Current", "Change"]
        sec = self._section(
            parent,
            f"Student improvement — {len(data)} student(s)")
        self._export_row(sec, hdrs, data)
        self._data_table(sec, hdrs, data, color_col=4)

    # ── Most improved subjects ────────────────────────────────
    def _tab_improved_subjects(self, parent, asmt_id,
                                class_id, asmt2_id):
        if not asmt2_id:
            info = ctk.CTkFrame(parent, fg_color=ACCENT_BG,
                                corner_radius=8)
            info.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(info,
                         text="Select a baseline assessment in 'Compare with' above.",
                         font=("", 12), text_color=ACCENT,
                         ).pack(padx=14, pady=10, anchor="w")
            return

        where  = "AND m.class_id=?" if class_id else ""
        p1 = (asmt_id,  class_id) if class_id else (asmt_id,)
        p2 = (asmt2_id, class_id) if class_id else (asmt2_id,)

        curr = {r["subject_id"]: r["mean"] for r in query(f"""
            SELECT subject_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY subject_id
        """, p1)}

        prev = {r["subject_id"]: r["mean"] for r in query(f"""
            SELECT subject_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY subject_id
        """, p2)}

        subj_names = {r["id"]: r["name"] for r in
                      query("SELECT id, name FROM subjects")}

        rows = []
        for sid, cur_m in curr.items():
            if sid in prev:
                diff = round(cur_m - prev[sid], 1)
                rows.append((subj_names.get(sid, "—"),
                              prev[sid], cur_m, diff))

        rows.sort(key=lambda x: x[3], reverse=True)

        if not rows:
            muted(parent, "No subjects found in both assessments."
                  ).pack(pady=24)
            return

        data = []
        for name, prev_m, cur_m, diff in rows:
            arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "—")
            data.append([name, f"{prev_m:.1f}%", f"{cur_m:.1f}%",
                          f"{arrow} {abs(diff):.1f}%"])

        hdrs = ["Subject", "Previous mean", "Current mean", "Change"]
        sec = self._section(parent, "Subject improvement")
        self._export_row(sec, hdrs, data)
        self._data_table(sec, hdrs, data, color_col=3)

        sec2 = self._section(parent, "Improvement chart")
        chart_data = [{"label": r[0], "value": r[3]} for r in rows]
        self._bar_chart(sec2, chart_data, "label", "value",
                        show_value=True)

    # ── Reusable widgets ──────────────────────────────────────
    def _section(self, parent, title, exportable_rows=None,
                 export_headers=None):
        c = card(parent)
        c.pack(fill="x", pady=(0, 14))
        hdr = ctk.CTkFrame(c, fg_color=ACCENT_BG, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text=title, font=("", 13, "bold"),
                     text_color=ACCENT).pack(
            anchor="w", padx=16, pady=8, side="left")
        return c

    def _export_row(self, parent, headers, rows):
        """Add export buttons to a section."""
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=12, pady=(8, 6))
        ghost_btn(f, "Print",
                  command=lambda: self._print_table(headers, rows),
                  width=80).pack(side="right", padx=(6, 0))
        ghost_btn(f, "Export PDF",
                  command=lambda: self._export_pdf(headers, rows),
                  width=100).pack(side="right", padx=(6, 0))
        ghost_btn(f, "Export Excel",
                  command=lambda: self._export_excel(headers, rows),
                  width=100).pack(side="right")

    def _export_excel(self, headers, rows):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="analytics_export.xlsx",
            title="Save Excel as")
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="4F46E5")
            cell.fill = PatternFill("solid", fgColor="EEF2FF")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(path)
        import os
        os.system(f'open "{path}"' if os.name != "nt"
                  else f'start "" "{path}"')

    def _print_table(self, headers, rows):
        """Generate PDF then send to printer."""
        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()
        self._export_pdf(headers, rows, path=tmp.name, silent=True)
        if os.name == "nt":
            os.startfile(tmp.name, "print")
        else:
            ret = os.system(f'lpr "{tmp.name}"')
            if ret != 0:
                os.system(f'open "{tmp.name}"')

    def _export_pdf(self, headers, rows, path=None, silent=False):
        from tkinter import filedialog
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from routes.settings import get_setting

        if not path:
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile="analytics_export.pdf",
                title="Save PDF as")
            if not path:
                return

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                 leftMargin=1.5*cm, rightMargin=1.5*cm,
                                 topMargin=1.5*cm, bottomMargin=1.5*cm)
        ACCENT_C  = colors.HexColor("#4F46E5")
        HDR_BG    = colors.HexColor("#EEF2FF")
        BORDER_C  = colors.HexColor("#E5E7EB")
        ALT       = colors.HexColor("#F9FAFB")

        school = get_setting("school_name", "GradeVault")
        title_style = ParagraphStyle("t", fontSize=13,
                                      fontName="Helvetica-Bold",
                                      textColor=ACCENT_C)
        story = [Paragraph(school, title_style), Spacer(1, 0.3*cm)]

        table_data = [headers] + [[str(c) for c in r] for r in rows]
        col_w = (26*cm) / len(headers)
        t = Table(table_data, colWidths=[col_w]*len(headers),
                  repeatRows=1)
        style = [
            ("BACKGROUND",    (0,0), (-1,0), HDR_BG),
            ("TEXTCOLOR",     (0,0), (-1,0), ACCENT_C),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("GRID",          (0,0), (-1,-1), 0.4, BORDER_C),
            ("ALIGN",         (0,0), (-1,-1), "LEFT"),
        ]
        for i in range(2, len(table_data), 2):
            style.append(("BACKGROUND", (0,i), (-1,i), ALT))
        t.setStyle(TableStyle(style))
        story.append(t)
        doc.build(story)
        import os
        if not silent:
            os.system(f'open "{path}"' if os.name != "nt"
                      else f'start "" "{path}"')

    def _data_table(self, parent, headers, rows,
                    highlight_top=0, color_col=None,
                    col_widths=None):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=12, pady=(6, 12))

        n     = len(headers)
        col_w = max(80, 660 // n)

        thead = ctk.CTkFrame(f, fg_color="#F3F4F6", corner_radius=6)
        thead.pack(fill="x", pady=(0, 2))
        for hi, h in enumerate(headers):
            w = col_widths[hi] if col_widths and hi < len(col_widths) else col_w
            ctk.CTkLabel(thead, text=h, font=("", 11, "bold"),
                         text_color=TEXT_MUTED, width=w,
                         anchor="w").pack(
                side="left", padx=(10, 0), pady=6)

        scroll = ctk.CTkScrollableFrame(
            f, fg_color="transparent",
            height=min(320, len(rows) * 36 + 10))
        scroll.pack(fill="x")

        for i, row in enumerate(rows):
            bg = SURFACE if i % 2 == 0 else "#FAFAFA"
            if highlight_top and i < highlight_top:
                bg = "#FEF9C3"
            r = ctk.CTkFrame(scroll, fg_color=bg,
                             corner_radius=0, height=34)
            r.pack(fill="x")
            r.pack_propagate(False)
            for j, cell in enumerate(row):
                color = TEXT
                if color_col is not None and j == color_col:
                    s = str(cell)
                    color = (SUCCESS if "▲" in s
                             else DANGER if "▼" in s
                             else TEXT_MUTED)
                if highlight_top and i < highlight_top and j == 0:
                    color = ACCENT
                w = col_widths[j] if col_widths and j < len(col_widths) else col_w
                ctk.CTkLabel(r, text=str(cell),
                             font=("", 11), text_color=color,
                             width=w, anchor="w").pack(
                    side="left", padx=(10, 0))

    def _bar_chart(self, parent, rows, label_key, value_key,
                   show_value=False):
        if not rows:
            return

        def safe_float(v):
            try:
                return float(str(v).replace("%", "").replace("▲", "")
                             .replace("▼", "").strip())
            except Exception:
                return 0.0

        if isinstance(rows[0], dict):
            labels = [str(r[label_key]) for r in rows]
            values = [safe_float(r[value_key]) for r in rows]
        else:
            labels = [str(r[label_key]) for r in rows]
            values = [safe_float(r[value_key]) for r in rows]

        if not values:
            return

        max_v   = max(abs(v) for v in values) or 1
        height  = 160
        bar_w   = max(28, min(60, 580 // len(labels)))
        chart_w = bar_w * len(labels) + 60

        import tkinter as tk
        canvas = tk.Canvas(parent,
                           width=max(chart_w, 580),
                           height=height + 60,
                           bg=SURFACE, highlightthickness=0)
        canvas.pack(padx=16, pady=(6, 12))

        baseline = height + 10
        for i, (lbl, val) in enumerate(zip(labels, values)):
            x     = 40 + i * bar_w + bar_w // 2
            bar_h = int(abs(val) / max_v * (height - 20))
            color = "#4F46E5" if val >= 0 else "#EF4444"
            y_top = baseline - bar_h

            canvas.create_rectangle(
                x - bar_w//2 + 3, y_top,
                x + bar_w//2 - 3, baseline,
                fill=color, outline="")

            if show_value:
                canvas.create_text(
                    x, y_top - 8,
                    text=f"{val:.1f}",
                    font=("Helvetica", 8), fill="#374151")

            short = lbl[:9] + "…" if len(lbl) > 10 else lbl
            canvas.create_text(
                x, baseline + 14,
                text=short, font=("Helvetica", 8),
                fill="#6B7280")

        canvas.create_line(40, baseline, max(chart_w, 580),
                           baseline, fill="#E5E7EB", width=1)
