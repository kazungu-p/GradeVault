"""
Analytics tab rendering methods — imported by AnalyticsPage.
Kept separate to reduce per-import load time.
"""
import customtkinter as ctk
from utils.theme import *
from db.connection import query, query_one


class AnalyticsTabsMixin:
    """Mix-in that adds all tab-rendering methods to AnalyticsPage."""
    def _tab_overview(self, parent, asmt_id, class_id, asmt2_id):
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        if class_id:
            stats = query_one("""
                SELECT COUNT(DISTINCT student_id) AS students,
                       ROUND(AVG(percentage), 1)  AS mean,
                       ROUND(MIN(percentage), 1)  AS min_pct,
                       ROUND(MAX(percentage), 1)  AS max_pct
                FROM marks_new
                WHERE assessment_id=? AND class_id=?
            """, (asmt_id, class_id)) or {}
        else:
            stats = query_one("""
                SELECT COUNT(DISTINCT student_id) AS students,
                       ROUND(AVG(percentage), 1)  AS mean,
                       ROUND(MIN(percentage), 1)  AS min_pct,
                       ROUND(MAX(percentage), 1)  AS max_pct
                FROM marks_new
                WHERE assessment_id=?
            """, (asmt_id,)) or {}

        mean     = stats.get("mean")    or 0
        students = stats.get("students") or 0
        min_pct  = stats.get("min_pct") or 0
        max_pct  = stats.get("max_pct") or 0

        if class_id:
            pass_count = (query_one("""
                SELECT COUNT(*) AS n FROM (
                    SELECT student_id, AVG(percentage) AS avg_pct
                    FROM marks_new WHERE assessment_id=? AND class_id=?
                    GROUP BY student_id
                ) WHERE avg_pct >= 50
            """, (asmt_id, class_id)) or {}).get("n", 0)
        else:
            pass_count = (query_one("""
                SELECT COUNT(*) AS n FROM (
                    SELECT student_id, AVG(percentage) AS avg_pct
                    FROM marks_new WHERE assessment_id=?
                    GROUP BY student_id
                ) WHERE avg_pct >= 50
            """, (asmt_id,)) or {}).get("n", 0)

        pass_rate = round(pass_count / students * 100, 1) if students else 0

        # Stat cards
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=(0, 16))
        row.columnconfigure((0, 1, 2, 3), weight=1, uniform="s")

        for col, (lbl, val, sub, color) in enumerate([
            ("Students analysed", str(students),    "with marks entered",   TEXT_MUTED),
            ("School mean",       f"{mean:.1f}",    "overall average",      ACCENT),
            ("Pass rate",         f"{pass_rate}",   "students with ≥ 50%",  SUCCESS if pass_rate >= 50 else DANGER),
            ("Score range",       f"{max_pct:.0f}",  f"Low: {min_pct:.0f}",   TEXT_MUTED),
        ]):
            c = ctk.CTkFrame(row, fg_color=SURFACE,
                             border_color=BORDER, border_width=1,
                             corner_radius=8)
            c.grid(row=0, column=col,
                   padx=(0, 10) if col < 3 else 0, sticky="ew")
            inner = ctk.CTkFrame(c, fg_color="transparent")
            inner.pack(padx=14, pady=12, fill="both")
            muted(inner, lbl, size=11).pack(anchor="w")
            label(inner, val, size=22, weight="bold").pack(
                anchor="w", pady=(2, 0))
            ctk.CTkLabel(inner, text=sub, font=("", 11),
                         text_color=color).pack(anchor="w")

        # Print / export overview button
        ov_btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        ov_btn_row.pack(fill="x", pady=(0, 8))
        _ov_stats = {"students": students, "mean": mean,
                     "pass_rate": pass_rate,
                     "max_pct": max_pct, "min_pct": min_pct}
        ghost_btn(ov_btn_row, "Print / Export overview",
                  command=lambda s=_ov_stats, a=asmt_id, c=class_id:
                      self._show_print_overview(a, c, s),
                  width=190).pack(side="right")

        # Grade distribution
        if class_id:
            grade_rows = query("""
                SELECT grade, COUNT(*) AS n
                FROM marks_new
                WHERE assessment_id=? AND class_id=?
                GROUP BY grade ORDER BY n DESC
            """, (asmt_id, class_id))
        else:
            grade_rows = query("""
                SELECT grade, COUNT(*) AS n
                FROM marks_new WHERE assessment_id=?
                GROUP BY grade ORDER BY n DESC
            """, (asmt_id,))

        if grade_rows:
            sec = self._section(parent, "Grade distribution")
            self._bar_chart(sec, grade_rows, "grade", "n",
                            show_value=True, chart_title="Grade Distribution")

        # Per-class performance
        cls_rows = query("""
            SELECT c.name || COALESCE(' '||c.stream,'') AS cls,
                   ROUND(AVG(m.percentage),1) AS mean,
                   COUNT(DISTINCT m.student_id) AS students
            FROM marks_new m
            JOIN classes c ON m.class_id=c.id
            WHERE m.assessment_id=?
            GROUP BY m.class_id ORDER BY mean DESC
        """, (asmt_id,))

        if cls_rows:
            sec2 = self._section(parent, "Performance by class")
            self._data_table(sec2,
                             ["Class", "Mean", "Students"],
                             [[r["cls"], f"{r['mean']:.1f}",
                               str(r["students"])]
                              for r in cls_rows])

    # ── Subject performance ───────────────────────────────────
    def _tab_subjects(self, parent, asmt_id, class_id, asmt2_id):
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        rows = query(f"""
            SELECT s.name AS subject,
                   ROUND(AVG(m.percentage),1) AS mean,
                   ROUND(MIN(m.percentage),1) AS min_pct,
                   ROUND(MAX(m.percentage),1) AS max_pct,
                   COUNT(*) AS entries
            FROM marks_new m
            JOIN subjects s ON m.subject_id=s.id
            WHERE m.assessment_id=? {where}
            GROUP BY m.subject_id ORDER BY mean DESC
        """, params)

        if not rows:
            muted(parent, "No marks found.").pack(pady=24)
            return

        sec = self._section(parent, "Subject performance — ranked by mean")
        self._bar_chart(sec, rows, "subject", "mean", show_value=True, chart_title="Subject Performance")

        tbl = [[str(i+1), r["subject"], f"{r['mean']:.1f}",
                 f"{r['min_pct']:.1f}", f"{r['max_pct']:.1f}",
                 str(r["entries"])]
                for i, r in enumerate(rows)]
        sec2 = self._section(parent, "Detail")
        self._export_row(sec2,
            ["#", "Subject", "Mean", "Min", "Max", "Count"], tbl)
        self._data_table(sec2,
            ["#", "Subject", "Mean", "Min", "Max", "Count"], tbl,
            col_widths=[40, 280, 90, 90, 90, 70])

    # ── Exam ranking ──────────────────────────────────────────
    def _tab_ranking(self, parent, asmt_id, class_id, asmt2_id):
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        rows = query(f"""
            SELECT st.id AS student_id, st.full_name, st.admission_number,
                   c.name || COALESCE(' '||c.stream,'') AS cls,
                   ROUND(AVG(m.percentage),1) AS mean
            FROM marks_new m
            JOIN students st ON m.student_id=st.id
            JOIN classes  c  ON m.class_id=c.id
            WHERE m.assessment_id=? {where}
            GROUP BY m.student_id ORDER BY mean DESC
        """, params)

        if not rows:
            muted(parent, "No marks found.").pack(pady=24)
            return

        from utils.grading import grade_from_percentage, detect_curriculum
        show_class = class_id is None

        data = []
        pos  = 1
        for i, r in enumerate(rows):
            if i > 0 and r["mean"] < rows[i-1]["mean"]:
                pos = i + 1
            cls_name   = r["cls"].split()[0] if r["cls"] else ""
            curriculum = detect_curriculum(cls_name)
            grade, _   = grade_from_percentage(r["mean"], curriculum)
            row_data = [str(pos), r["full_name"], r["admission_number"]]
            if show_class:
                row_data.append(r["cls"])
            row_data += [f"{r['mean']:.1f}", grade]
            data.append(row_data)

        headers = ["#", "Full name", "Adm. No."]
        if show_class:
            headers.append("Class")
        headers += ["Mean", "Grade"]

        widths = [40, 220, 120]
        if show_class:
            widths.append(150)
        widths += [70, 70]

        sec = self._section(parent, f"Exam ranking — {len(data)} student(s)")
        self._export_row(sec, headers, data)
        self._paginated_table(sec, headers, data,
                              highlight_top=3, col_widths=widths,
                              page_size=20)


    def _tab_improved_students(self, parent, asmt_id,
                                class_id, asmt2_id):
        if not asmt2_id:
            info = ctk.CTkFrame(parent, fg_color=ACCENT_BG,
                                corner_radius=8)
            info.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(info,
                         text="Select a baseline assessment in 'Compare with' above.",
                         font=("", 12), text_color=ACCENT,
                         ).pack(padx=14, pady=10, anchor="w")
            return

        where  = "AND class_id=?" if class_id else ""
        p1 = (asmt_id,  class_id) if class_id else (asmt_id,)
        p2 = (asmt2_id, class_id) if class_id else (asmt2_id,)

        curr = {r["student_id"]: r["mean"] for r in query(f"""
            SELECT student_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY student_id
        """, p1)}

        prev = {r["student_id"]: r["mean"] for r in query(f"""
            SELECT student_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY student_id
        """, p2)}

        improvements = []
        for sid, cur_m in curr.items():
            if sid in prev:
                diff = round(cur_m - prev[sid], 1)
                improvements.append((sid, prev[sid], cur_m, diff))

        if not improvements:
            muted(parent, "No students found in both assessments."
                  ).pack(pady=24)
            return

        improvements.sort(key=lambda x: x[3], reverse=True)

        id_map = {r["id"]: r for r in query("""
            SELECT s.id, s.full_name, s.admission_number,
                   c.name || COALESCE(' ' || c.stream, '') AS cls
            FROM students s
            JOIN classes c ON s.class_id = c.id
        """)}

        data = []
        for sid, prev_m, cur_m, diff in improvements:
            st = id_map.get(sid, {})
            arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "—")
            data.append([st.get("full_name", "—"),
                          st.get("admission_number", "—"),
                          st.get("cls", "—").strip(),
                          f"{prev_m:.1f}", f"{cur_m:.1f}",
                          f"{arrow} {abs(diff):.1f}"])

        hdrs = ["Full name", "Adm. No.", "Class", "Previous", "Current", "Change"]
        sec = self._section(
            parent,
            f"Student improvement — {len(data)} student(s)")
        self._export_row(sec, hdrs, data)
        self._data_table(sec, hdrs, data, color_col=5,
                         col_widths=[200, 110, 130, 80, 80, 90])

        # ── Most improved student per subject — bulk fetch ────
        where2 = "AND class_id=?" if class_id else ""
        p_base = (class_id,) if class_id else ()

        # Two bulk queries instead of 2*N
        curr_all = query(f"""
            SELECT subject_id, student_id,
                   ROUND(AVG(percentage),1) AS mean
            FROM marks_new
            WHERE assessment_id=? {where2}
            GROUP BY subject_id, student_id
        """, (asmt_id,) + p_base)
        prev_all = query(f"""
            SELECT subject_id, student_id,
                   ROUND(AVG(percentage),1) AS mean
            FROM marks_new
            WHERE assessment_id=? {where2}
            GROUP BY subject_id, student_id
        """, (asmt2_id,) + p_base)

        from collections import defaultdict
        curr_by_subj = defaultdict(dict)
        prev_by_subj = defaultdict(dict)
        for r in curr_all:
            curr_by_subj[r["subject_id"]][r["student_id"]] = r["mean"]
        for r in prev_all:
            prev_by_subj[r["subject_id"]][r["student_id"]] = r["mean"]

        subj_name_map = {r["id"]: r["name"]
                         for r in query("SELECT id, name FROM subjects")}
        per_subj_data = []
        for subj_id, c_map in curr_by_subj.items():
            p_map = prev_by_subj.get(subj_id, {})
            best_diff, best_sid, best_prev, best_curr = None, None, 0, 0
            for st_id, cur_m in c_map.items():
                if st_id in p_map:
                    diff = round(cur_m - p_map[st_id], 1)
                    if best_diff is None or diff > best_diff:
                        best_diff = diff
                        best_sid  = st_id
                        best_prev = p_map[st_id]
                        best_curr = cur_m
            if best_sid is not None and best_diff is not None:
                st    = id_map.get(best_sid, {})
                arrow = "▲" if best_diff > 0 else ("▼" if best_diff < 0 else "—")
                per_subj_data.append([
                    subj_name_map.get(subj_id, "—"),
                    st.get("full_name", "—"),
                    st.get("admission_number", "—"),
                    st.get("cls", "—").strip(),
                    f"{best_prev:.1f}", f"{best_curr:.1f}",
                    f"{arrow} {abs(best_diff):.1f}",
                ])
        per_subj_data.sort(key=lambda r: r[0])  # sort by subject name

        if per_subj_data:
            hdrs2 = ["Subject", "Most improved student", "Adm. No.",
                     "Class", "Previous", "Current", "Change"]
            sec2 = self._section(parent, "Most improved student per subject")
            self._export_row(sec2, hdrs2, per_subj_data)
            self._data_table(sec2, hdrs2, per_subj_data, color_col=6,
                             col_widths=[150, 180, 100, 120, 65, 65, 75])

    # ── Most improved subjects ────────────────────────────────
    def _tab_improved_subjects(self, parent, asmt_id,
                                class_id, asmt2_id):
        if not asmt2_id:
            info = ctk.CTkFrame(parent, fg_color=ACCENT_BG,
                                corner_radius=8)
            info.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(info,
                         text="Select a baseline assessment in 'Compare with' above.",
                         font=("", 12), text_color=ACCENT,
                         ).pack(padx=14, pady=10, anchor="w")
            return

        where  = "AND class_id=?" if class_id else ""
        p1 = (asmt_id,  class_id) if class_id else (asmt_id,)
        p2 = (asmt2_id, class_id) if class_id else (asmt2_id,)

        curr = {r["subject_id"]: r["mean"] for r in query(f"""
            SELECT subject_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY subject_id
        """, p1)}

        prev = {r["subject_id"]: r["mean"] for r in query(f"""
            SELECT subject_id, ROUND(AVG(percentage),1) AS mean
            FROM marks_new WHERE assessment_id=? {where}
            GROUP BY subject_id
        """, p2)}

        subj_names = {r["id"]: r["name"] for r in
                      query("SELECT id, name FROM subjects")}

        rows = []
        for sid, cur_m in curr.items():
            if sid in prev:
                diff = round(cur_m - prev[sid], 1)
                rows.append((subj_names.get(sid, "—"),
                              prev[sid], cur_m, diff))

        rows.sort(key=lambda x: x[3], reverse=True)

        if not rows:
            muted(parent, "No subjects found in both assessments."
                  ).pack(pady=24)
            return

        data = []
        for name, prev_m, cur_m, diff in rows:
            arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "—")
            data.append([name, f"{prev_m:.1f}", f"{cur_m:.1f}",
                          f"{arrow} {abs(diff):.1f}"])

        hdrs = ["Subject", "Previous mean", "Current mean", "Change"]
        sec = self._section(parent, "Subject improvement")
        self._export_row(sec, hdrs, data)
        self._data_table(sec, hdrs, data, color_col=3)

        sec2 = self._section(parent, "Improvement chart")
        chart_data = [{"label": r[0], "value": r[3]} for r in rows]
        self._bar_chart(sec2, chart_data, "label", "value",
                        show_value=True, chart_title="Subject Improvement")

    # ── Top per subject ──────────────────────────────────────
    def _tab_top_per_subject(self, parent, asmt_id, class_id, asmt2_id):
        """Best student per subject + best performing subjects."""
        where  = "AND m.class_id=?" if class_id else ""
        params = (asmt_id, class_id) if class_id else (asmt_id,)

        # Best student per subject
        subj_rows = query(f"""
            SELECT s.name AS subject,
                   st.full_name, st.admission_number,
                   c.name || COALESCE(' ' || c.stream, '') AS cls,
                   ROUND(m.percentage,1) AS pct
            FROM marks_new m
            JOIN subjects s  ON m.subject_id = s.id
            JOIN students st ON m.student_id = st.id
            JOIN classes  c  ON m.class_id   = c.id
            WHERE m.assessment_id=? {where}
              AND m.percentage = (
                  SELECT MAX(m2.percentage)
                  FROM marks_new m2
                  WHERE m2.assessment_id = m.assessment_id
                    AND m2.subject_id = m.subject_id
                    {"AND m2.class_id=?" if class_id else ""}
              )
            ORDER BY s.name
        """, params + ((class_id,) if class_id else ()))

        if subj_rows:
            sec = self._section(parent, "Top student per subject")
            hdrs = ["Subject", "Top student", "Adm. No.", "Class", "Score"]
            data = [[r["subject"], r["full_name"],
                     r["admission_number"],
                     r["cls"].strip(), f"{r['pct']:.1f}"]
                    for r in subj_rows]
            self._export_row(sec, hdrs, data)
            self._data_table(sec, hdrs, data,
                             col_widths=[180, 200, 110, 130, 80])

        # Best performing subjects (ranked by mean)
        best_rows = query(f"""
            SELECT s.name AS subject,
                   ROUND(AVG(m.percentage),1) AS mean,
                   ROUND(MAX(m.percentage),1) AS top_score,
                   COUNT(*) AS entries
            FROM marks_new m
            JOIN subjects s ON m.subject_id = s.id
            WHERE m.assessment_id=? {where}
            GROUP BY m.subject_id
            ORDER BY mean DESC
        """, params)

        if best_rows:
            sec2 = self._section(parent, "Best performing subjects")
            hdrs2 = ["#", "Subject", "Mean", "Top score", "Students"]
            data2 = [[str(i+1), r["subject"], f"{r['mean']:.1f}",
                      f"{r['top_score']:.1f}", str(r["entries"])]
                     for i, r in enumerate(best_rows)]
            self._export_row(sec2, hdrs2, data2)
            self._data_table(sec2, hdrs2, data2,
                             col_widths=[40, 260, 80, 80, 80])
            self._bar_chart(sec2, best_rows, "subject", "mean",
                            show_value=True)

        if not subj_rows and not best_rows:
            muted(parent, "No marks found.").pack(pady=24)

    # ── Combined assessments ──────────────────────────────────
    def _tab_combined(self, parent, asmt_id, class_id, asmt2_id):
        """Aggregate analysis across multiple assessments."""
        ids = [i for i, _ in self._multi_selected]
        if not ids:
            info = ctk.CTkFrame(parent, fg_color=ACCENT_BG, corner_radius=8)
            info.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(
                info,
                text="Use 'Add assessment' above to select two or more assessments to combine.",
                font=("", 12), text_color=ACCENT,
            ).pack(padx=14, pady=10, anchor="w")
            return

        names = [n for _, n in self._multi_selected]
        where  = "AND m.class_id=?" if class_id else ""
        p_base = (class_id,) if class_id else ()

        placeholders = ",".join("?" * len(ids))

        # Per-subject mean across all selected assessments
        subj_rows = query(f"""
            SELECT s.name AS subject,
                   ROUND(AVG(m.percentage),1) AS combined_mean,
                   ROUND(MIN(m.percentage),1) AS min_pct,
                   ROUND(MAX(m.percentage),1) AS max_pct,
                   COUNT(DISTINCT m.assessment_id) AS asmt_count
            FROM marks_new m
            JOIN subjects s ON m.subject_id = s.id
            WHERE m.assessment_id IN ({placeholders}) {where}
            GROUP BY m.subject_id
            ORDER BY combined_mean DESC
        """, tuple(ids) + p_base)

        if subj_rows:
            sec = self._section(
                parent,
                f"Subject means — combined ({', '.join(names)})")
            hdrs = ["Subject", "Combined mean", "Min", "Max", "# Assessments"]
            data = [[r["subject"], f"{r['combined_mean']:.1f}",
                     f"{r['min_pct']:.1f}", f"{r['max_pct']:.1f}",
                     str(r["asmt_count"])]
                    for r in subj_rows]
            self._export_row(sec, hdrs, data)
            self._data_table(sec, hdrs, data,
                             col_widths=[220, 120, 80, 80, 100])
            self._bar_chart(sec, subj_rows, "subject", "combined_mean",
                            show_value=True)

        # Per-student overall mean across all selected assessments
        student_rows = query(f"""
            SELECT st.full_name, st.admission_number,
                   c.name || COALESCE(' '||c.stream,'') AS cls,
                   ROUND(AVG(m.percentage),1) AS combined_mean,
                   COUNT(DISTINCT m.assessment_id) AS asmt_count
            FROM marks_new m
            JOIN students st ON m.student_id = st.id
            JOIN classes  c  ON m.class_id = c.id
            WHERE m.assessment_id IN ({placeholders}) {where}
            GROUP BY m.student_id
            ORDER BY combined_mean DESC
        """, tuple(ids) + p_base)

        if student_rows:
            show_cls = class_id is None
            sec2 = self._section(
                parent,
                f"Student combined ranking — {len(student_rows)} student(s)")
            hdrs2 = ["#", "Full name", "Adm. No."]
            widths2 = [40, 220, 120]
            if show_cls:
                hdrs2.append("Class")
                widths2.append(140)
            hdrs2 += ["Combined mean", "Assessments"]
            widths2 += [120, 100]
            data2 = []
            for i, r in enumerate(student_rows):
                row = [str(i+1), r["full_name"], r["admission_number"]]
                if show_cls:
                    row.append(r["cls"])
                row += [f"{r['combined_mean']:.1f}", str(r["asmt_count"])]
                data2.append(row)
            self._export_row(sec2, hdrs2, data2)
            self._data_table(sec2, hdrs2, data2,
                             highlight_top=3, col_widths=widths2)

        if not subj_rows and not student_rows:
            muted(parent, "No marks found for the selected assessments."
                  ).pack(pady=24)


    def _paginated_table(self, parent, headers, rows,
                         highlight_top=0, color_col=None,
                         col_widths=None, page_size=20):
        """Data table with built-in pagination for large datasets."""
        state = {"page": 0}
        total_pages = max(1, (len(rows) + page_size - 1) // page_size)

        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="x", padx=12, pady=(6, 12))

        # Page nav bar
        nav = ctk.CTkFrame(container, fg_color="transparent")
        nav.pack(fill="x", pady=(0, 6))
        page_lbl = muted(nav, f"Page 1 / {total_pages}  ·  {len(rows)} students")
        page_lbl.pack(side="left")
        next_btn = ghost_btn(nav, "Next →", command=None, width=80)
        next_btn.pack(side="right")
        prev_btn = ghost_btn(nav, "← Prev", command=None, width=80)
        prev_btn.pack(side="right", padx=(0, 6))

        table_host = ctk.CTkFrame(container, fg_color="transparent")
        table_host.pack(fill="x")

        def _render():
            for w in table_host.winfo_children():
                w.destroy()
            p     = state["page"]
            chunk = rows[p * page_size: (p + 1) * page_size]
            self._data_table(table_host, headers, chunk,
                             highlight_top=highlight_top if p == 0 else 0,
                             color_col=color_col, col_widths=col_widths)
            page_lbl.configure(
                text=f"Page {p+1} / {total_pages}  ·  {len(rows)} students")
            prev_btn.configure(state="normal" if p > 0 else "disabled")
            next_btn.configure(
                state="normal" if p < total_pages - 1 else "disabled")

        def _prev():
            if state["page"] > 0:
                state["page"] -= 1
                _render()

        def _next():
            if state["page"] < total_pages - 1:
                state["page"] += 1
                _render()

        prev_btn.configure(command=_prev)
        next_btn.configure(command=_next)
        prev_btn.configure(state="disabled")
        next_btn.configure(
            state="normal" if total_pages > 1 else "disabled")
        _render()

    # ── Reusable widgets ──────────────────────────────────────
    def _section(self, parent, title):
        c = card(parent)
        c.pack(fill="x", pady=(0, 14))
        hdr = ctk.CTkFrame(c, fg_color=ACCENT_BG, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text=title, font=("", 13, "bold"),
                     text_color=ACCENT, anchor="w").pack(
            fill="x", padx=16, pady=8)
        return c

    def _export_row(self, parent, headers, rows):
        """Add export buttons to a section."""
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=12, pady=(8, 6))
        ghost_btn(f, "Print",
                  command=lambda: self._print_table(headers, rows),
                  width=80).pack(side="right", padx=(6, 0))
        ghost_btn(f, "Export PDF",
                  command=lambda: self._export_pdf(headers, rows),
                  width=100).pack(side="right", padx=(6, 0))
        ghost_btn(f, "Export Excel",
                  command=lambda: self._export_excel(headers, rows),
                  width=100).pack(side="right")

    def _download_chart(self, canvas, title, width, height):
        """
        Export chart to PDF using only ReportLab — no Ghostscript, no EPS.
        Re-draws the chart natively using ReportLab's canvas API.
        """
        from tkinter import filedialog
        from routes.settings import get_setting
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors as rlc
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas as rl_canvas
        from pathlib import Path
        import os

        safe = title.replace(" ", "_").replace("/", "-")
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF file", "*.pdf")],
            initialfile=f"chart_{safe}.pdf",
            title="Save chart as")
        if not path:
            return

        school  = get_setting("school_name", "GradeVault")
        motto   = get_setting("school_motto", "")
        contact = get_setting("school_contact", "")
        logo_p  = get_setting("school_logo", "")
        AC_HEX  = "#4F46E5"

        PW, PH = landscape(A4)
        c = rl_canvas.Canvas(path, pagesize=(PW, PH))

        # ── Header ───────────────────────────────────────────────
        HDR_H = 80   # header band height in points

        c.setFillColor(rlc.HexColor("#EEF2FF"))
        c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)

        logo_drawn = False
        if logo_p and Path(logo_p).exists():
            try:
                from reportlab.lib.utils import ImageReader
                img_r = ImageReader(logo_p)
                iw, ih = img_r.getSize()
                ratio  = iw / ih
                # Fill full width (with 1cm margins each side)
                lw = PW - 2 * cm
                lh = lw / ratio
                if lh > HDR_H - 8:          # cap height to header band
                    lh = HDR_H - 8
                    lw = lh * ratio
                c.drawImage(img_r,
                            PW / 2 - lw / 2,
                            PH - HDR_H + (HDR_H - lh) / 2,
                            width=lw, height=lh,
                            preserveAspectRatio=True, mask="auto")
                logo_drawn = True
            except Exception:
                pass

        if not logo_drawn:
            # Text header — school name centred, motto and contact below
            c.setFillColor(rlc.HexColor(AC_HEX))
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(PW / 2, PH - 28, school)
            if motto:
                c.setFillColor(rlc.HexColor("#6B7280"))
                c.setFont("Helvetica-Oblique", 9)
                c.drawCentredString(PW / 2, PH - 42, motto)
            if contact:
                c.setFillColor(rlc.HexColor("#6B7280"))
                c.setFont("Helvetica", 8)
                c.drawCentredString(PW / 2, PH - 54, contact)

        # Divider
        c.setStrokeColor(rlc.HexColor(AC_HEX))
        c.setLineWidth(1)
        c.line(1.5*cm, PH - HDR_H - 2, PW - 1.5*cm, PH - HDR_H - 2)

        # Chart title
        c.setFillColor(rlc.HexColor(AC_HEX))
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(PW / 2, PH - HDR_H - 18, title)

        # ── Re-draw chart from tkinter canvas items ───────────────
        # Tkinter: origin top-left, y grows down
        # ReportLab: origin bottom-left, y grows up
        margin_x = 1.5 * cm
        chart_top = PH - HDR_H - 28     # PDF y of chart top
        chart_bottom = 1.5 * cm
        chart_left = margin_x
        chart_right = PW - margin_x
        avail_w = chart_right - chart_left
        avail_h = chart_top - chart_bottom

        # Scale factors from tkinter canvas coords to PDF coords
        sx = avail_w / max(width, 1)
        sy = avail_h / max(height, 1)

        def tk_to_pdf(tx, ty):
            """Convert tkinter (tx, ty) to PDF coords."""
            px = chart_left + tx * sx
            py = chart_top  - ty * sy   # flip Y
            return px, py

        # Iterate tkinter canvas items and reproduce in PDF
        for item_id in canvas.find_all():
            try:
                itype  = canvas.type(item_id)
                coords = canvas.coords(item_id)
                opts   = {k: canvas.itemcget(item_id, k)
                          for k in canvas.itemconfig(item_id)}

                def hex_or_none(val):
                    v = str(val).strip()
                    if not v or v in ("", "None"): return None
                    try:
                        return rlc.HexColor(v) if v.startswith("#")                                else rlc.toColor(v)
                    except Exception:
                        return None

                fill   = hex_or_none(opts.get("fill", ""))
                outline= hex_or_none(opts.get("outline", ""))
                width_v= float(opts.get("width", 1) or 1)

                if itype == "line" and len(coords) >= 4:
                    pts = [(coords[i], coords[i+1])
                           for i in range(0, len(coords)-1, 2)]
                    p = c.beginPath()
                    x0, y0 = tk_to_pdf(pts[0][0], pts[0][1])
                    p.moveTo(x0, y0)
                    for tx, ty in pts[1:]:
                        px, py = tk_to_pdf(tx, ty)
                        p.lineTo(px, py)
                    c.setStrokeColor(fill or rlc.black)
                    c.setLineWidth(width_v * min(sx, sy))
                    c.drawPath(p, stroke=1, fill=0)

                elif itype == "oval" and len(coords) == 4:
                    x1,y1,x2,y2 = coords
                    cx, cy = tk_to_pdf((x1+x2)/2, (y1+y2)/2)
                    rx = (x2-x1)/2 * sx
                    ry = (y2-y1)/2 * sy
                    c.setFillColor(fill or rlc.white)
                    c.setStrokeColor(outline or rlc.black)
                    c.setLineWidth(width_v * min(sx, sy))
                    c.ellipse(cx-rx, cy-ry, cx+rx, cy+ry,
                              fill=1, stroke=1 if outline else 0)

                elif itype == "rectangle" and len(coords) == 4:
                    x1,y1,x2,y2 = coords
                    px1,py1 = tk_to_pdf(x1, y2)
                    px2,py2 = tk_to_pdf(x2, y1)
                    if fill:
                        c.setFillColor(fill)
                        c.rect(px1, py1, px2-px1, py2-py1, fill=1, stroke=0)

                elif itype == "text":
                    if len(coords) >= 2:
                        tx, ty = coords[0], coords[1]
                        px, py = tk_to_pdf(tx, ty)
                        text   = opts.get("text", "")
                        fsize  = 7
                        try:
                            fspec = opts.get("font", "")
                            if fspec:
                                parts = str(fspec).split()
                                for p_item in parts:
                                    try:
                                        fsize = float(p_item)
                                        break
                                    except ValueError:
                                        pass
                        except Exception:
                            pass
                        anchor = opts.get("anchor", "center")
                        c.setFillColor(fill or rlc.black)
                        c.setFont("Helvetica", max(5, fsize * min(sx,sy)))
                        if "e" in anchor:
                            c.drawRightString(px, py, str(text))
                        elif "w" in anchor:
                            c.drawString(px, py, str(text))
                        else:
                            c.drawCentredString(px, py, str(text))
            except Exception:
                pass

        c.save()
        os.system(f'open "{path}"' if os.name != "nt"
                  else f'start "" "{path}"')

    def _export_excel(self, headers, rows):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="analytics_export.xlsx",
            title="Save Excel as")
        if not path:
            return
        from routes.settings import get_setting as gs2
        school  = gs2("school_name", "GradeVault")
        motto   = gs2("school_motto", "")
        contact = gs2("school_contact", "")
        term_label = ""
        try:
            from routes.terms import get_current_term
            t = get_current_term()
            if t:
                term_label = f"Term {t['term']}, {t['year']}"
        except Exception:
            pass

        wb = openpyxl.Workbook()
        ws = wb.active
        # School header rows
        ws.append([school])
        ws["A1"].font = Font(bold=True, size=14, color="4F46E5")
        if motto:
            ws.append([motto])
            ws[f"A2"].font = Font(italic=True, size=10, color="6B7280")
        if contact:
            ws.append([contact])
        if term_label:
            ws.append([term_label])
        ws.append([])  # blank row
        hdr_row = ws.max_row + 1
        ws.append(headers)
        for cell in ws[hdr_row]:
            cell.font = Font(bold=True, color="4F46E5")
            cell.fill = PatternFill("solid", fgColor="EEF2FF")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(path)
        import os
        os.system(f'open "{path}"' if os.name != "nt"
                  else f'start "" "{path}"')

    def _print_table(self, headers, rows):
        """Generate PDF then open printer selection dialog."""
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()
        self._export_pdf(headers, rows, path=tmp.name, silent=True)
        from pages.reports import PrintDialog
        PrintDialog(self, tmp.name, status_label=None)

    def _show_print_overview(self, asmt_id, class_id, stats):
        """Build a simple overview PDF and open printer dialog."""
        import tempfile
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from routes.settings import get_setting

        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()

        doc = SimpleDocTemplate(tmp.name, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        ACCENT_C = colors.HexColor("#4F46E5")
        HDR_BG   = colors.HexColor("#EEF2FF")
        BORDER_C = colors.HexColor("#E5E7EB")

        school  = get_setting("school_name", "GradeVault")
        motto   = get_setting("school_motto", "")
        contact = get_setting("school_contact", "")
        logo_p  = get_setting("school_logo", "")
        term_label = ""
        try:
            from routes.terms import get_current_term
            t2 = get_current_term()
            if t2:
                term_label = f"Term {t2['term']}, {t2['year']}"
        except Exception:
            pass

        title_s = ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold",
                                  textColor=ACCENT_C, alignment=1)
        motto_s = ParagraphStyle("mo", fontSize=9, fontName="Helvetica-Oblique",
                                  textColor=colors.HexColor("#6B7280"), alignment=1)
        sub_s   = ParagraphStyle("s", fontSize=10, fontName="Helvetica",
                                  textColor=colors.HexColor("#374151"), alignment=1)
        from reportlab.platypus import HRFlowable as HRF2
        from reportlab.platypus import Image as _Img2
        from pathlib import Path as _P2
        story = []
        logo_shown2 = False
        if logo_p and _P2(logo_p).exists():
            try:
                img = _Img2(logo_p)
                ratio = img.imageWidth / img.imageHeight
                w = 17*cm
                h = w / ratio
                if h > 4*cm:
                    h = 4*cm; w = h * ratio
                img.drawWidth, img.drawHeight = w, h
                img.hAlign = "CENTER"
                story.append(img)
                logo_shown2 = True
            except Exception:
                pass
        if not logo_shown2:
            story.append(Paragraph(school, title_s))
            if motto:   story.append(Paragraph(motto, motto_s))
            if contact: story.append(Paragraph(contact, motto_s))
        if term_label:
            story.append(Paragraph(term_label, sub_s))
        story.append(Paragraph("School Overview Report", sub_s))
        story.append(Spacer(1, 0.2*cm))
        story.append(HRF2(width="100%", thickness=0.5,
                          color=ACCENT_C, spaceAfter=6))
        data = [
            ["Metric", "Value"],
            ["Students Analysed",   str(stats.get("students", 0))],
            ["School Mean",         f"{stats.get('mean', 0):.1f}"],
            ["Pass Rate",           f"{stats.get('pass_rate', 0):.1f}"],
            ["Highest Score",       f"{stats.get('max_pct', 0):.0f}"],
            ["Lowest Score",        f"{stats.get('min_pct', 0):.0f}"],
        ]
        t = Table(data, colWidths=[8*cm, 8*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), HDR_BG),
            ("TEXTCOLOR",     (0,0), (-1,0), ACCENT_C),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME",      (0,1), (0,-1), "Helvetica-Bold"),
            ("FONTNAME",      (1,1), (1,-1), "Helvetica"),
            ("FONTSIZE",      (0,0), (-1,-1), 11),
            ("TOPPADDING",    (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING",   (0,0), (-1,-1), 12),
            ("GRID",          (0,0), (-1,-1), 0.5, BORDER_C),
            ("ROWBACKGROUNDS",(0,1), (-1,-1),
             [colors.white, colors.HexColor("#F9FAFB")]),
        ]))
        story.append(t)
        doc.build(story)

        from pages.reports import PrintDialog
        PrintDialog(self, tmp.name, status_label=None)

    def _export_pdf(self, headers, rows, path=None, silent=False):
        from tkinter import filedialog
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from routes.settings import get_setting

        if not path:
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile="analytics_export.pdf",
                title="Save PDF as")
            if not path:
                return

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                 leftMargin=1.5*cm, rightMargin=1.5*cm,
                                 topMargin=1.5*cm, bottomMargin=1.5*cm)
        ACCENT_C  = colors.HexColor("#4F46E5")
        HDR_BG    = colors.HexColor("#EEF2FF")
        BORDER_C  = colors.HexColor("#E5E7EB")
        ALT       = colors.HexColor("#F9FAFB")

        school  = get_setting("school_name", "GradeVault")
        motto   = get_setting("school_motto", "")
        contact = get_setting("school_contact", "")
        logo_p  = get_setting("school_logo", "")
        term_label = ""
        try:
            from routes.terms import get_current_term
            t = get_current_term()
            if t:
                term_label = f"Term {t['term']}, {t['year']}"
        except Exception:
            pass

        title_style = ParagraphStyle("t", fontSize=14,
                                      fontName="Helvetica-Bold",
                                      textColor=ACCENT_C, alignment=1)
        motto_style = ParagraphStyle("m", fontSize=9,
                                      fontName="Helvetica-Oblique",
                                      textColor=colors.HexColor("#6B7280"),
                                      alignment=1)
        sub_style   = ParagraphStyle("s", fontSize=10,
                                      fontName="Helvetica",
                                      textColor=colors.HexColor("#374151"),
                                      alignment=1)
        from reportlab.platypus import HRFlowable, Image as _Img4
        from pathlib import Path as _P4
        story = []
        logo_shown = False
        if logo_p and _P4(logo_p).exists():
            try:
                img = _Img4(logo_p)
                ratio = img.imageWidth / img.imageHeight
                w = 25*cm
                h = w / ratio
                if h > 4*cm:
                    h = 4*cm; w = h * ratio
                img.drawWidth, img.drawHeight = w, h
                img.hAlign = "CENTER"
                story.append(img)
                logo_shown = True
            except Exception:
                pass
        if not logo_shown:
            # No logo — show text header
            story.append(Paragraph(school, title_style))
            if motto:   story.append(Paragraph(motto, motto_style))
            if contact: story.append(Paragraph(contact, motto_style))
        if term_label:
            story.append(Paragraph(term_label, sub_style))
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=ACCENT_C, spaceAfter=6))

        table_data = [headers] + [[str(c) for c in r] for r in rows]
        col_w = (26*cm) / len(headers)
        t = Table(table_data, colWidths=[col_w]*len(headers),
                  repeatRows=1)
        style = [
            ("BACKGROUND",    (0,0), (-1,0), HDR_BG),
            ("TEXTCOLOR",     (0,0), (-1,0), ACCENT_C),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("GRID",          (0,0), (-1,-1), 0.4, BORDER_C),
            ("ALIGN",         (0,0), (-1,-1), "LEFT"),
        ]
        for i in range(2, len(table_data), 2):
            style.append(("BACKGROUND", (0,i), (-1,i), ALT))
        t.setStyle(TableStyle(style))
        story.append(t)
        doc.build(story)
        import os
        if not silent:
            os.system(f'open "{path}"' if os.name != "nt"
                      else f'start "" "{path}"')

    def _data_table(self, parent, headers, rows,
                    highlight_top=0, color_col=None,
                    col_widths=None):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=12, pady=(6, 12))

        n     = len(headers)
        col_w = max(80, 660 // n)

        thead = ctk.CTkFrame(f, fg_color="#F3F4F6", corner_radius=6)
        thead.pack(fill="x", pady=(0, 2))
        for hi, h in enumerate(headers):
            w = col_widths[hi] if col_widths and hi < len(col_widths) else col_w
            ctk.CTkLabel(thead, text=h, font=("", 11, "bold"),
                         text_color=TEXT_MUTED, width=w,
                         anchor="w").pack(
                side="left", padx=(10, 0), pady=6)

        needed_h = len(rows) * 36 + 10
        scroll = ctk.CTkScrollableFrame(
            f, fg_color="transparent",
            height=min(320, needed_h),
            scrollbar_button_color=BG,
            scrollbar_button_hover_color=BG)
        scroll.pack(fill="x")
        # Hide scrollbar when content fits
        try:
            if needed_h <= 320:
                scroll._scrollbar.grid_remove()
            else:
                scroll._scrollbar.configure(width=6)
        except Exception:
            pass

        for i, row in enumerate(rows):
            bg = SURFACE if i % 2 == 0 else "#FAFAFA"
            if highlight_top and i < highlight_top:
                bg = "#FEF9C3"
            r = ctk.CTkFrame(scroll, fg_color=bg,
                             corner_radius=0, height=34)
            r.pack(fill="x")
            r.pack_propagate(False)
            for j, cell in enumerate(row):
                color = TEXT
                if color_col is not None and j == color_col:
                    s = str(cell)
                    color = (SUCCESS if "▲" in s
                             else DANGER if "▼" in s
                             else TEXT_MUTED)
                if highlight_top and i < highlight_top and j == 0:
                    color = ACCENT
                w = col_widths[j] if col_widths and j < len(col_widths) else col_w
                ctk.CTkLabel(r, text=str(cell),
                             font=("", 11), text_color=color,
                             width=w, anchor="w").pack(
                    side="left", padx=(10, 0))

    def _bar_chart(self, parent, rows, label_key, value_key,
                   show_value=False, chart_title="Chart"):
        if not rows:
            return

        def safe_float(v):
            try:
                return float(str(v).replace("%", "").replace("▲", "")
                             .replace("▼", "").strip())
            except Exception:
                return 0.0

        if isinstance(rows[0], dict):
            labels = [str(r[label_key]) for r in rows]
            values = [safe_float(r[value_key]) for r in rows]
        else:
            labels = [str(r[label_key]) for r in rows]
            values = [safe_float(r[value_key]) for r in rows]

        if not values:
            return

        max_v   = max(abs(v) for v in values) or 1
        height  = 160
        bar_w   = max(28, min(60, 580 // len(labels)))
        chart_w = max(bar_w * len(labels) + 60, 580)
        total_h = height + 60

        import tkinter as tk
        canvas = tk.Canvas(parent,
                           width=chart_w,
                           height=total_h,
                           bg=SURFACE, highlightthickness=0)
        canvas.pack(padx=16, pady=(6, 4))

        baseline = height + 10
        for i, (lbl, val) in enumerate(zip(labels, values)):
            x     = 40 + i * bar_w + bar_w // 2
            bar_h = int(abs(val) / max_v * (height - 20))
            color = "#4F46E5" if val >= 0 else "#EF4444"
            y_top = baseline - bar_h

            canvas.create_rectangle(
                x - bar_w//2 + 3, y_top,
                x + bar_w//2 - 3, baseline,
                fill=color, outline="")

            if show_value:
                canvas.create_text(
                    x, y_top - 8,
                    text=f"{val:.1f}",
                    font=("Helvetica", 8), fill="#374151")

            short = lbl[:9] + "…" if len(lbl) > 10 else lbl
            canvas.create_text(
                x, baseline + 14,
                text=short, font=("Helvetica", 8),
                fill="#6B7280")

        canvas.create_line(40, baseline, chart_w,
                           baseline, fill="#E5E7EB", width=1)

        # Download button below chart
        dl_row = ctk.CTkFrame(parent, fg_color="transparent")
        dl_row.pack(fill="x", padx=16, pady=(0, 10))
        ghost_btn(dl_row, "↓ Download chart",
                  command=lambda c=canvas, t=chart_title, w=chart_w, h=total_h:
                      self._download_chart(c, t, w, h),
                  width=140).pack(side="right")
